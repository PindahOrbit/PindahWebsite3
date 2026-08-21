#!/usr/bin/env python3
"""Merge HDX/MoH health facilities (with GPS) into the providers database."""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

import build_providers_db as core

ROOT = Path(__file__).resolve().parent
REPO = ROOT.parent
OUT = ROOT / "output"
HDX = ROOT / "raw" / "other" / "health-facility-list-with-geo-codes.xlsx"
TODAY = date.today().isoformat()

OWNERSHIP = {
    "1": "Government",
    "2": "Government",
    "3": "Mission",
    "4": "Private",
    "5": "Local Authority / Rural",
    "6": "NGO",
    "7": "Other",
}


def load_hdx() -> list[dict[str, str]]:
    rows = []
    wb = openpyxl.load_workbook(HDX, data_only=True)
    ws = wb.active
    headers = [core.clean(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    core.add_source(
        "https://data.humdata.org/dataset/zimbabwe-health-facilities/resource/c60fac4d-b3fb-48a7-9652-7d7e44b34a89",
        "HDX / MoH Zimbabwe Health Facility List (with geo-codes)",
        "Humanitarian Data Exchange facility master list with latitude/longitude",
    )
    for raw in ws.iter_rows(min_row=2, values_only=True):
        data = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
        name = core.clean(data.get("NAMEOFFACI"))
        if not name or name.startswith("#"):
            continue
        fac_type = core.clean(data.get("TYPE_EDITE") or data.get("TYPEOFFACI"))
        ownership = OWNERSHIP.get(str(core.clean(data.get("OWNERSHIP"))), core.clean(data.get("OWNERSHIP")))
        province = core.clean(data.get("Province")).replace("Matebeleland", "Matabeleland")
        district = core.clean(data.get("DISTRICT"))
        lat = data.get("LATITUDE")
        lon = data.get("LONGITUDE")
        r = core.empty_row()
        # Many names are short place names; append facility type for clarity when needed
        display = name
        if fac_type and fac_type.lower() not in name.lower():
            display = f"{name} {fac_type}".strip()
        r["Provider Name"] = display
        r["City"] = core.normalize_city(district)
        r["Province"] = province.title() if province else core.infer_province(district)
        r["Address"] = f"{district}, {province}".strip(", ")
        if lat not in (None, "", 0, "0") and lon not in (None, "", 0, "0"):
            try:
                r["Latitude"] = str(float(lat))
                r["Longitude"] = str(float(lon))
            except Exception:
                pass
        low = fac_type.lower()
        if "hospital" in low:
            if ownership == "Mission":
                r["Provider Type"] = "Mission Hospital"
            elif ownership == "Government":
                r["Provider Type"] = "Government Hospital"
            elif ownership == "Private":
                r["Provider Type"] = "Private Hospital"
            else:
                r["Provider Type"] = "Hospital"
        elif "clinic" in low or "polyclinic" in low:
            r["Provider Type"] = "Clinic"
        elif "rural health" in low:
            r["Provider Type"] = "Clinic"
        else:
            r["Provider Type"] = fac_type or "Healthcare Facility"
        r["Organisation"] = ownership
        r["Source URL"] = (
            "https://data.humdata.org/dataset/1f291458-c999-434a-8160-e13fceaac957/resource/"
            "c60fac4d-b3fb-48a7-9652-7d7e44b34a89/download/health-facility-list-with-geo-codes.xlsx"
        )
        r["Source Title"] = "HDX MoH Health Facility List"
        r["Last Verified"] = TODAY
        notes = [f"Facility type: {fac_type}" if fac_type else "", f"Ownership code/source: {ownership}"]
        if data.get("UPDATED"):
            notes.append(f"Source updated field: {data.get('UPDATED')}")
        r["Notes"] = "; ".join(n for n in notes if n)
        rows.append(r)
    print("HDX facilities:", len(rows))
    return rows


def load_current() -> list[dict[str, str]]:
    path = REPO / "providers.csv"
    with path.open(encoding="utf-8") as f:
        return [{c: row.get(c, "") for c in core.COLUMNS} for row in csv.DictReader(f)]


def export(rows: list[dict[str, str]]) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    # reload prior sources.csv and append
    prev_sources = REPO / "sources.csv"
    if prev_sources.exists():
        with prev_sources.open(encoding="utf-8") as f:
            for row in csv.DictReader(f):
                core.add_source(row.get("URL", ""), row.get("Title", ""), row.get("Notes", ""))

    providers_csv = OUT / "providers.csv"
    with providers_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=core.COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in core.COLUMNS})
    (REPO / "providers.csv").write_bytes(providers_csv.read_bytes())

    sources_csv = OUT / "sources.csv"
    with sources_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["URL", "Title", "Accessed", "Notes"])
        w.writeheader()
        for s in sorted(core.sources, key=lambda x: x["URL"]):
            w.writerow(s)
    (REPO / "sources.csv").write_bytes(sources_csv.read_bytes())

    wb = Workbook()
    ws = wb.active
    ws.title = "Providers"
    ws.append(core.COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c, "") for c in core.COLUMNS])
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(OUT / "Zimbabwe_Healthcare_Providers.xlsx")
    wb.save(REPO / "Zimbabwe_Healthcare_Providers.xlsx")

    counts = core.classify_counts(rows)
    with_gps = sum(1 for r in rows if r.get("Latitude") and r.get("Longitude"))
    pdf_count = len(list((ROOT / "raw" / "psmas_pdfs").glob("*.pdf"))) + (
        1 if (ROOT / "raw" / "other" / "CellMed-Service-Provider-Directory.pdf").exists() else 0
    )
    summary = f"""Zimbabwe Healthcare Providers Database
Generated: {TODAY}

Number of providers collected: {len(rows)}
Number with AHFoZ numbers: {counts.get('with_ahfoz', 0)}
Number with registration numbers: {counts.get('with_reg', 0)}
Number of hospitals: {max(counts.get('hospitals', 0), counts.get('hospitals_name', 0))}
Number of doctors: {counts.get('doctors', 0)}
Number of pharmacies: {max(counts.get('pharmacies', 0), counts.get('pharmacies_name', 0))}
Number of laboratories: {max(counts.get('laboratories', 0), counts.get('laboratories_name', 0))}
Number with GPS coordinates: {with_gps}
Number of duplicates removed: {core.stats.get('duplicates_removed', 0)}
Websites searched: {len(core.sources)}
PDFs processed: {pdf_count}
Total rows: {len(rows)}

Major sources:
1. PSMAS Provider Network PDFs (July 2026) — main public source of AHFoZ / practice numbers
2. PSMAS Excel provider list (repository)
3. Health Professions Authority registered facilities
4. MDPCZ Public Register (full ~4,650 practitioners with registration numbers)
5. HDX / Ministry of Health health facility list with geo-codes
6. CellMed Service Provider Directory PDF
7. ZACH membership directory
8. PSMAS website listings + directory crawls (CIMAS Blue Zone, Hello Doctor, medical.co.zw, AHFoZ)

Important limitations:
- AHFoZ does not publish a complete public payee directory. AHFoZ numbers here come from medical-aid network publications (mainly PSMAS). Blank AHFoZ means not found publicly — not fabricated.
- Alliance Health's older provider PDF is no longer linked on their current website.
- Some HDX/MoH facility records are historical master-list entries; names/types retained with source attribution.
"""
    (OUT / "summary.txt").write_text(summary, encoding="utf-8")
    (REPO / "summary.txt").write_text(summary, encoding="utf-8")
    print(summary)


def main() -> None:
    current = load_current()
    print("Current rows:", len(current))
    hdx = load_hdx()
    merged = core.merge_records(current + hdx)
    print("Merged rows:", len(merged))
    export(merged)


if __name__ == "__main__":
    main()
