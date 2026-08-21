#!/usr/bin/env python3
"""Enrich providers DB with MDPCZ full register + CellMed PDF, then rebuild exports."""

from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import date
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

import build_providers_db as core

ROOT = Path(__file__).resolve().parent
RAW_OTHER = ROOT / "raw" / "other"
OUT = ROOT / "output"
REPO = ROOT.parent
TODAY = date.today().isoformat()
MDPCZ_JSON = RAW_OTHER / "mdpcz_full.json"
CELLMED_PDF = RAW_OTHER / "CellMed-Service-Provider-Directory.pdf"


def load_mdpcz_full() -> list[dict[str, str]]:
    rows = []
    if not MDPCZ_JSON.exists():
        print("MDPCZ full JSON missing")
        return rows
    data = json.loads(MDPCZ_JSON.read_text(encoding="utf-8"))
    core.add_source(
        "https://www.mdpcz.co.zw/public_register",
        "MDPCZ Public Register (full scrape)",
        f"{len(data)} practitioners via Livewire pagination",
    )
    for item in data:
        name = core.clean(item.get("Name"))
        if not name:
            continue
        specialty = core.clean(item.get("Specialty"))
        qual = core.clean(item.get("Qualification"))
        reg = core.clean(item.get("RegistrationNumber") or item.get("Registration Number"))
        r = core.empty_row()
        r["Provider Name"] = name
        r["Registration Number"] = reg
        r["Speciality"] = specialty
        low = (specialty + " " + qual).lower()
        if "dental" in low or specialty.lower().startswith("dental"):
            r["Provider Type"] = "Dentist"
        elif "intern" in specialty.lower():
            r["Provider Type"] = "Doctor"
            r["Speciality"] = specialty or "Intern"
        else:
            r["Provider Type"] = "Doctor"
        r["Source URL"] = "https://www.mdpcz.co.zw/public_register"
        r["Source Title"] = "MDPCZ Public Register"
        r["Last Verified"] = TODAY
        notes = []
        gender = core.clean(item.get("Gender"))
        if gender:
            notes.append(f"Gender: {gender}")
        if qual:
            notes.append(f"Qualification: {qual}")
        r["Notes"] = "; ".join(notes)
        rows.append(r)
    print(f"MDPCZ full: {len(rows)}")
    return rows


def parse_cellmed_pdf() -> list[dict[str, str]]:
    rows = []
    if not CELLMED_PDF.exists():
        return rows
    source_url = "https://zimmedicover.com/wp-content/uploads/2024/12/CellMed-Service-Provider-Directory.pdf"
    core.add_source(source_url, "CellMed Service Provider Directory PDF", "No AHFoZ numbers in this directory")
    section = ""
    with pdfplumber.open(CELLMED_PDF) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                up = core.clean(line).upper()
                if up and len(up) < 60 and any(
                    k in up
                    for k in [
                        "GENERAL PRACTITION",
                        "SPECIALIST",
                        "PHARMAC",
                        "DENT",
                        "LABORATOR",
                        "HOSPITAL",
                        "OPTIC",
                        "PHYSIO",
                        "RADIOLOGY",
                        "CLINIC",
                    ]
                ):
                    section = up
            for table in page.extract_tables() or []:
                for raw in table:
                    cells = [core.clean((c or "").replace("\n", " ")) for c in raw]
                    compact = [c for c in cells if c]
                    if not compact:
                        continue
                    joined = " ".join(compact).upper()
                    if "PROVIDER" in joined and "ADDRESS" in joined:
                        continue
                    if len(compact) <= 2 and not re.search(r"\d{6,}", joined):
                        if any(k in joined for k in ["GENERAL", "PHARMAC", "DENT", "LAB", "HOSPITAL", "SPECIAL"]):
                            section = joined
                        continue
                    # patterns: town, provider, address, contact  OR provider sparse
                    town = provider = address = contact = ""
                    if len(compact) >= 4:
                        town, provider, address, contact = compact[0], compact[1], compact[2], compact[3]
                    elif len(compact) == 3:
                        # could be town, provider, contact OR provider, address, contact
                        if re.search(r"\d{6,}", compact[2]):
                            if re.search(r"\d", compact[1]) and not re.search(r"\d{6,}", compact[1]):
                                town, provider, contact = compact[0], compact[1], compact[2]
                                address = ""
                            else:
                                provider, address, contact = compact[0], compact[1], compact[2]
                        else:
                            town, provider, address = compact
                    else:
                        continue
                    # Heuristic swap if "provider" looks like town-only and address missing
                    if provider and not address and len(compact) >= 4:
                        pass
                    name = provider or ""
                    if not name or name.upper() in {"PROVIDER", "GENERAL PRACTITIONERS"}:
                        continue
                    if name.upper() == compact[0].upper() and len(compact) >= 4:
                        # sometimes first is town
                        pass
                    r = core.empty_row()
                    r["Provider Name"] = name
                    r["Address"] = address
                    r["City"] = core.normalize_city(town if town.lower() not in {"harare/bulawayo/gweru/rusape/kwekwe"} else "")
                    if "/" in (town or "") and not r["City"]:
                        r["Notes"] = f"Towns: {town}"
                        r["City"] = ""
                    r["Province"] = core.infer_province(r["City"])
                    phone = core.normalize_phone(contact)
                    if phone.startswith("0") and len(re.sub(r"\D", "", phone)) >= 9:
                        r["Mobile"] = phone
                    else:
                        r["Phone"] = phone
                    ptype, spec = core.infer_type_from_name(name, section)
                    r["Provider Type"] = ptype or "Healthcare Provider"
                    r["Speciality"] = spec
                    r["Medical Aid Accepted"] = "CellMed"
                    r["Source URL"] = source_url
                    r["Source Title"] = "CellMed Service Provider Directory"
                    r["Last Verified"] = TODAY
                    note = f"Section: {section}" if section else ""
                    if r.get("Notes"):
                        note = (note + " | " + r["Notes"]).strip(" |")
                    r["Notes"] = note
                    rows.append(r)
    print(f"CellMed: {len(rows)}")
    return rows


def load_existing_providers_csv() -> list[dict[str, str]]:
    path = REPO / "providers.csv"
    if not path.exists():
        return []
    with path.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def export(rows: list[dict[str, str]]) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    # csv
    providers_csv = OUT / "providers.csv"
    with providers_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=core.COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in core.COLUMNS})
    (REPO / "providers.csv").write_bytes(providers_csv.read_bytes())

    sources_csv = OUT / "sources.csv"
    with sources_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["URL", "Title", "Accessed", "Notes"])
        w.writeheader()
        for s in sorted(core.sources, key=lambda x: x["URL"]):
            w.writerow(s)
    # ensure key sources present
    (REPO / "sources.csv").write_bytes(sources_csv.read_bytes())

    wb = Workbook()
    ws = wb.active
    ws.title = "Providers"
    ws.append(core.COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c, "") for c in core.COLUMNS])
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(OUT / "Zimbabwe_Healthcare_Providers.xlsx")
    wb.save(REPO / "Zimbabwe_Healthcare_Providers.xlsx")

    counts = core.classify_counts(rows)
    pdfs = list((ROOT / "raw" / "psmas_pdfs").glob("*.pdf"))
    if CELLMED_PDF.exists():
        pdfs.append(CELLMED_PDF)
    summary = f"""Zimbabwe Healthcare Providers Database
Generated: {TODAY}

Number of providers collected: {len(rows)}
Number with AHFoZ numbers: {counts.get('with_ahfoz', 0)}
Number with registration numbers: {counts.get('with_reg', 0)}
Number of hospitals: {max(counts.get('hospitals', 0), counts.get('hospitals_name', 0))}
Number of doctors: {counts.get('doctors', 0)}
Number of pharmacies: {max(counts.get('pharmacies', 0), counts.get('pharmacies_name', 0))}
Number of laboratories: {max(counts.get('laboratories', 0), counts.get('laboratories_name', 0))}
Number of duplicates removed: {core.stats.get('duplicates_removed', 0)}
Websites searched: {len(core.sources)}
PDFs processed: {len(pdfs)}
Total rows: {len(rows)}

Major sources used:
- PSMAS provincial + consolidated provider network PDFs (July 2026) — primary public AHFoZ/payee numbers
- Repository PSMAS Excel (AHFoZ numbers)
- Health Professions Authority registered facilities listing
- MDPCZ public register full scrape ({len(json.loads(MDPCZ_JSON.read_text(encoding='utf-8'))) if MDPCZ_JSON.exists() else 0} practitioners with registration numbers)
- CellMed Service Provider Directory PDF
- ZACH membership directory
- PSMAS website listings
- CIMAS Blue Zone, Hello Doctor, Zimbabwe Medical Directory, AHFoZ site (searched)

Notes:
- AHFoZ does not publish a complete public provider payee directory. AHFoZ numbers included here are those published in medical-aid network lists (chiefly PSMAS).
- Where an AHFoZ number was not publicly available it was left blank (not invented).
- HPA website intermittently blocked automated access; facilities were captured from an accessible public listing snapshot.
- Alliance Health historical provider PDF was no longer linked on their current site at collection time.
"""
    (OUT / "summary.txt").write_text(summary, encoding="utf-8")
    (REPO / "summary.txt").write_text(summary, encoding="utf-8")
    print(summary)


def main() -> None:
    # Start from previous build rows + fresh MDPCZ/CellMed
    all_rows: list[dict[str, str]] = []

    # Re-run core collectors for freshness of AHFoZ-bearing sources
    print("Collecting core sources...")
    all_rows.extend(core.parse_existing_xlsx())
    all_rows.extend(core.parse_all_psmas_pdfs())
    all_rows.extend(core.parse_hpa_markdown(core.HPA_CACHE))
    all_rows.extend(core.fetch_zach_members())
    all_rows.extend(core.fetch_psmas_wp_listings())
    all_rows.extend(core.fetch_additional_directories())

    print("Adding MDPCZ full + CellMed...")
    all_rows.extend(load_mdpcz_full())
    all_rows.extend(parse_cellmed_pdf())

    print("Raw before dedupe:", len(all_rows))
    merged = core.merge_records(all_rows)
    print("After dedupe:", len(merged))
    export(merged)
    (OUT / "stats.json").write_text(json.dumps(dict(core.stats), indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
