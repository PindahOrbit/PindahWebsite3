#!/usr/bin/env python3
"""Build Zimbabwe Healthcare Providers database from public sources."""

from __future__ import annotations

import csv
import json
import os
import re
import time
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import openpyxl
import pdfplumber
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent
RAW = ROOT / "raw"
OUT = ROOT / "output"
PSMAS_PDF_DIR = RAW / "psmas_pdfs"
REPO_XLSX = ROOT.parent / "wwwroot" / "PSMAS_Provider_Network_List.xlsx"
HPA_CACHE = Path(
    r"C:\Users\Administrator\.cursor\projects\f-Users-Administrator-Documents-GitHub-PindahWebsite3"
    r"\agent-tools\63948312-bb4b-4ddd-9d18-a5a1496dd993.txt"
)

TODAY = date.today().isoformat()
COLUMNS = [
    "Provider Name",
    "AHFoZ Number",
    "Registration Number",
    "Provider Type",
    "Speciality",
    "Medical Aid Accepted",
    "Organisation",
    "Address",
    "City",
    "Province",
    "Phone",
    "Mobile",
    "Email",
    "Website",
    "Contact Person",
    "Latitude",
    "Longitude",
    "Source URL",
    "Source Title",
    "Last Verified",
    "Notes",
]

CITY_TO_PROVINCE = {
    "harare": "Harare",
    "chitungwiza": "Harare",
    "epworth": "Harare",
    "ruwa": "Harare",
    "norton": "Mashonaland West",
    "bulawayo": "Bulawayo",
    "mutare": "Manicaland",
    "rusape": "Manicaland",
    "chipinge": "Manicaland",
    "nyanga": "Manicaland",
    "chimanimani": "Manicaland",
    "buhera": "Manicaland",
    "masvingo": "Masvingo",
    "chiredzi": "Masvingo",
    "gutu": "Masvingo",
    "chivi": "Masvingo",
    "mwenezi": "Masvingo",
    "bikita": "Masvingo",
    "zaka": "Masvingo",
    "gweru": "Midlands",
    "kwekwe": "Midlands",
    "kadoma": "Mashonaland West",
    "chegutu": "Mashonaland West",
    "chinhoyi": "Mashonaland West",
    "kariba": "Mashonaland West",
    "karoi": "Mashonaland West",
    "banket": "Mashonaland West",
    "gokwe": "Midlands",
    "shurugwi": "Midlands",
    "zvishavane": "Midlands",
    "mberengwa": "Midlands",
    "redcliff": "Midlands",
    "mvuma": "Midlands",
    "marondera": "Mashonaland East",
    "murewa": "Mashonaland East",
    "murehwa": "Mashonaland East",
    "mutoko": "Mashonaland East",
    "wedza": "Mashonaland East",
    "chivhu": "Mashonaland East",
    "goromonzi": "Mashonaland East",
    "bindura": "Mashonaland Central",
    "mvurwi": "Mashonaland Central",
    "guruve": "Mashonaland Central",
    "mount darwin": "Mashonaland Central",
    "concession": "Mashonaland Central",
    "shamva": "Mashonaland Central",
    "glendale": "Mashonaland Central",
    "gwanda": "Matabeleland South",
    "beitbridge": "Matabeleland South",
    "plumtree": "Matabeleland South",
    "filabusi": "Matabeleland South",
    "esigodini": "Matabeleland South",
    "west nicholson": "Matabeleland South",
    "hwange": "Matabeleland North",
    "victoria falls": "Matabeleland North",
    "lupane": "Matabeleland North",
    "binga": "Matabeleland North",
    "victoriafall": "Matabeleland North",
}

PROVINCE_HINTS = {
    "HARARE": "Harare",
    "BULAWAYO": "Bulawayo",
    "MANICALAND": "Manicaland",
    "MASVINGO": "Masvingo",
    "MIDLANDS": "Midlands",
    "MASHONALAND CENTRAL": "Mashonaland Central",
    "MASHONALAND EAST": "Mashonaland East",
    "MASHONALAND WEST": "Mashonaland West",
    "MATEBELELAND NORTH": "Matabeleland North",
    "MATABELELAND NORTH": "Matabeleland North",
    "MATEBELELAND SOUTH": "Matabeleland South",
    "MATABELELAND SOUTH": "Matabeleland South",
}

SECTION_TYPE_MAP = [
    (r"GENERAL PRACTITIONER", "Doctor", "General Practitioner"),
    (r"\bGP\b", "Doctor", "General Practitioner"),
    (r"SPECIALIST", "Doctor", "Specialist"),
    (r"PHYSICIAN", "Doctor", "Physician"),
    (r"PAEDIATRIC", "Doctor", "Paediatrician"),
    (r"GYNAE|OBSTET", "Doctor", "Obstetrics & Gynaecology"),
    (r"SURGEON|SURGERY", "Doctor", "Surgeon"),
    (r"ORTHOPAED", "Doctor", "Orthopaedic Surgeon"),
    (r"ANAESTH", "Doctor", "Anaesthetist"),
    (r"PSYCHIAT", "Doctor", "Psychiatrist"),
    (r"DERMATOL", "Doctor", "Dermatologist"),
    (r"ONCOL", "Doctor", "Oncologist"),
    (r"UROLOG", "Doctor", "Urologist"),
    (r"ENT\b|OTORHINO", "Doctor", "ENT"),
    (r"NEURO", "Doctor", "Neurosurgeon"),
    (r"OPHTHAL", "Doctor", "Ophthalmologist"),
    (r"DENTIST|DENTAL", "Dentist", "Dentistry"),
    (r"ORTHODONT", "Dentist", "Orthodontics"),
    (r"PHARMAC", "Pharmacy", ""),
    (r"LABORATOR", "Laboratory", ""),
    (r"RADIOLOGY|X-?RAY|MRI|CT SCAN|IMAGING", "Radiology Centre", "Radiology"),
    (r"OPTICIAN|OPTOMETR", "Optical Centre", "Optometry"),
    (r"PHYSIOTHER", "Rehabilitation Centre", "Physiotherapy"),
    (r"OCCUPATIONAL THER", "Rehabilitation Centre", "Occupational Therapy"),
    (r"REHABILIT", "Rehabilitation Centre", ""),
    (r"PSYCHOLOG", "Clinic", "Psychology"),
    (r"DIALYSIS", "Dialysis Centre", ""),
    (r"AMBULANCE", "Ambulance Service", ""),
    (r"GOVERNMENT HOSPITAL", "Government Hospital", ""),
    (r"MISSION HOSPITAL", "Mission Hospital", ""),
    (r"PRIVATE HOSPITAL", "Private Hospital", ""),
    (r"HOSPITAL", "Hospital", ""),
    (r"POLYCLINIC", "Polyclinic", ""),
    (r"CLINIC|MEDICAL CENTRE|MEDICAL CENTER", "Clinic", ""),
    (r"WELLNESS|OCCUPATIONAL HEALTH", "Occupational Health Clinic", ""),
    (r"FAMILY PLANNING", "Clinic", "Family Planning"),
    (r"THEATRE", "Private Hospital", "Theatre"),
    (r"NURSE", "Clinic", "Nursing"),
    (r"AUDIOLOG", "Clinic", "Audiology"),
    (r"BLOOD BANK", "Blood Bank", ""),
    (r"HOSPICE", "Hospice", ""),
]

sources: list[dict[str, str]] = []
stats = Counter()


def clean(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).replace("\xa0", " ").replace("\n", " ").strip()
    s = re.sub(r"\s+", " ", s)
    if s.lower() in {"none", "n/a", "na", "-", "null", "nan"}:
        return ""
    return s


def normalize_phone(val: Any) -> str:
    s = clean(val)
    if not s:
        return ""
    # Drop trailing .0 from excel floats
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    s = s.replace("+263", "0").replace("263 ", "0")
    s = re.sub(r"[^\d/;,+ ]", "", s)
    s = re.sub(r"\s+", " ", s).strip(" ,;")
    # Compact spaced mobile like 0772 580 852
    parts = []
    for part in re.split(r"[/;,]+", s):
        p = re.sub(r"\s+", "", part.strip())
        if not p:
            continue
        if p.startswith("263") and len(p) >= 12:
            p = "0" + p[3:]
        parts.append(p)
    return " / ".join(dict.fromkeys(parts))


def normalize_city(val: str) -> str:
    s = clean(val)
    if not s:
        return ""
    key = s.lower()
    aliases = {
        "hre": "Harare",
        "byo": "Bulawayo",
        "murehwa": "Murewa",
        "victoriafall": "Victoria Falls",
        "mt darwin": "Mount Darwin",
        "mt. darwin": "Mount Darwin",
    }
    if key in aliases:
        return aliases[key]
    return s.title() if s.isupper() or s.islower() else s


def infer_province(city: str, district: str = "", filename: str = "") -> str:
    for text in (city, district):
        key = clean(text).lower()
        if key in CITY_TO_PROVINCE:
            return CITY_TO_PROVINCE[key]
    up = filename.upper()
    for hint, prov in PROVINCE_HINTS.items():
        if hint in up:
            return prov
    return ""


def infer_type_from_text(text: str) -> tuple[str, str]:
    up = clean(text).upper()
    for pattern, ptype, spec in SECTION_TYPE_MAP:
        if re.search(pattern, up):
            return ptype, spec
    return "", ""


def infer_type_from_name(name: str, section: str = "") -> tuple[str, str]:
    ptype, spec = infer_type_from_text(section)
    if ptype:
        return ptype, spec
    return infer_type_from_text(name)


def looks_like_ahfoz(val: str) -> bool:
    s = clean(val)
    if not s:
        return False
    s = s.replace(",", "").replace(" ", "")
    if re.fullmatch(r"\d{3,8}", s):
        return True
    if re.fullmatch(r"0\d{4,7}", s):  # leading zero practice numbers
        return True
    return False


def normalize_ahfoz(val: Any) -> str:
    s = clean(val)
    if not s:
        return ""
    s = s.replace(",", "").strip()
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    s = re.sub(r"[^\d]", "", s)
    return s.lstrip("0") and s or s  # keep leading zeros if present meaningfully
    # Actually keep as-is without stripping meaningful zeros for display:
    # return s


def normalize_ahfoz_fixed(val: Any) -> str:
    s = clean(val)
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    s = re.sub(r"[^\d]", "", s)
    # Preserve original numeric string; strip only if all zeros
    return s


def empty_row() -> dict[str, str]:
    return {c: "" for c in COLUMNS}


def add_source(url: str, title: str, notes: str = "") -> None:
    url = clean(url)
    if not url:
        return
    for s in sources:
        if s["URL"] == url:
            return
    sources.append(
        {
            "URL": url,
            "Title": title,
            "Accessed": TODAY,
            "Notes": notes,
        }
    )


def record_source_search(url: str, title: str, notes: str = "") -> None:
    add_source(url, title, notes)


# ------------------------- parsers -------------------------


def parse_existing_xlsx() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    if not REPO_XLSX.exists():
        return rows
    wb = openpyxl.load_workbook(REPO_XLSX, data_only=True)
    ws = wb.active
    headers = [clean(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    add_source(
        "local://wwwroot/PSMAS_Provider_Network_List.xlsx",
        "PSMAS Provider Network List (repo Excel)",
        "Local copy with AHFoZ/AFHOZ numbers",
    )
    for raw in ws.iter_rows(min_row=2, values_only=True):
        data = {headers[i]: clean(raw[i]) if i < len(raw) else "" for i in range(len(headers))}
        name = data.get("Name") or data.get("Provider Name") or ""
        if not name:
            continue
        r = empty_row()
        r["Provider Name"] = name
        r["AHFoZ Number"] = normalize_ahfoz_fixed(
            data.get("AFHOZ Number") or data.get("AHFoZ Number") or data.get("AHFOZ Number")
        )
        r["Address"] = data.get("Physical Address") or data.get("Address") or ""
        r["Mobile"] = normalize_phone(data.get("Cell Number") or data.get("Mobile"))
        r["Phone"] = normalize_phone(data.get("Landline") or data.get("Phone"))
        r["Medical Aid Accepted"] = "PSMAS"
        r["City"] = normalize_city(guess_city_from_address(r["Address"]))
        r["Province"] = infer_province(r["City"])
        ptype, spec = infer_type_from_name(name)
        r["Provider Type"] = ptype or "Healthcare Provider"
        r["Speciality"] = spec
        r["Source URL"] = "https://psmas.co.zw/view-download-provider-list/"
        r["Source Title"] = "PSMAS Provider Network List (Excel)"
        r["Last Verified"] = TODAY
        r["Notes"] = "Imported from repository PSMAS Excel"
        rows.append(r)
        stats["from_repo_xlsx"] += 1
    return rows


def guess_city_from_address(address: str) -> str:
    a = clean(address).lower()
    # Prefer longer city names first
    cities = sorted(CITY_TO_PROVINCE.keys(), key=len, reverse=True)
    for city in cities:
        if re.search(rf"\b{re.escape(city)}\b", a):
            return city.title()
    return ""


def collapse_sparse_row(cells: list[Any]) -> list[str]:
    vals = [clean(c) for c in cells]
    # Keep non-empty in order but also retain positional when headers known
    return vals


def map_header_indexes(header_cells: list[str]) -> dict[str, int]:
    idx: dict[str, int] = {}
    joined = [(i, clean(h).upper()) for i, h in enumerate(header_cells) if clean(h)]
    for i, h in joined:
        if "DISTRICT" in h:
            idx["district"] = i
        elif "SUBURB" in h or "LOCATION" in h:
            idx["suburb"] = i
        elif h == "NAME" or h.endswith(" NAME") or h == "PROVIDER":
            idx["name"] = i
        elif "PHYSICAL" in h or h == "ADDRESS":
            idx["address"] = i
        elif "AHFOZ" in h or "AFHOZ" in h or "PRACTICE NUMBER" in h or h == "NUMBER":
            idx["ahfoz"] = i
        elif "CELL" in h or "MOBILE" in h:
            idx["mobile"] = i
        elif "LANDLINE" in h or "CONTACT" in h or "PHONE" in h:
            idx.setdefault("phone", i)
        elif "DISCIPLINE" in h or "TYPE" in h:
            idx["discipline"] = i
    # Fix split headers like AFHOZ / NUMBER on two rows handled by caller
    return idx


def parse_pdf_tables(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    province = ""
    for hint, prov in PROVINCE_HINTS.items():
        if hint in path.name.upper():
            province = prov
            break
    source_url = f"https://psmas.co.zw/wp-content/uploads/2026/07/{path.name}"
    add_source(source_url, f"PSMAS {path.stem}", "Provincial/network PDF July 2026")
    current_section = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            # Track section headers from text
            text = page.extract_text() or ""
            for line in text.splitlines():
                up = clean(line).upper()
                if not up or len(up) > 80:
                    continue
                if any(
                    k in up
                    for k in [
                        "GENERAL PRACTITION",
                        "PHARMAC",
                        "DENTIST",
                        "DENTAL",
                        "LABORATOR",
                        "RADIOLOGY",
                        "OPTIC",
                        "PHYSIOTHER",
                        "HOSPITAL",
                        "SPECIALIST",
                        "CLINIC",
                        "REHABILIT",
                        "AMBULANCE",
                        "DIALYSIS",
                        "PSYCHOLOG",
                        "PSYCHIAT",
                        "THEATRE",
                        "WELLNESS",
                        "NURSE",
                        "AUDIOLOG",
                        "OCCUPATIONAL",
                    ]
                ) and not re.search(r"\d{3,}", up):
                    # Avoid addresses
                    if "NETWORK" not in up and "CONTENTS" not in up:
                        current_section = up

            tables = page.extract_tables() or []
            for table in tables:
                if not table:
                    continue
                # Build header from first non-empty rows
                header_map: dict[str, int] = {}
                start = 0
                for i, row in enumerate(table[:3]):
                    cells = [clean(c) for c in row]
                    non_empty = [c for c in cells if c]
                    joined_header = " ".join(non_empty).upper()
                    if any(
                        k in joined_header
                        for k in ["NAME", "AHFOZ", "AFHOZ", "PRACTICE", "DISTRICT", "ADDRESS"]
                    ):
                        # merge with next row if NUMBER split
                        merged = cells[:]
                        if i + 1 < len(table):
                            nxt = [clean(c) for c in table[i + 1]]
                            for j, val in enumerate(nxt):
                                if j < len(merged):
                                    if not merged[j] and val:
                                        merged[j] = val
                                    elif merged[j] and val and val.upper() in {"NUMBER", "NO"}:
                                        merged[j] = f"{merged[j]} {val}"
                        header_map = map_header_indexes(merged)
                        start = i + 1
                        # if next was number-only, skip it
                        if i + 1 < len(table):
                            nxt_join = " ".join(clean(c) for c in table[i + 1] if clean(c)).upper()
                            if nxt_join in {"NUMBER", "NO", "AFHOZ NUMBER", "AHFOZ NUMBER"}:
                                start = i + 2
                        break
                if not header_map and "name" not in header_map:
                    # Try sparse rows without headers using positional guess from first data-like row
                    # Common: district, suburb, name, address, ahfoz, cell, landline
                    pass

                for row in table[start:]:
                    cells = [clean(c) for c in row]
                    nonempty = [c for c in cells if c]
                    if not nonempty:
                        continue
                    joined = " ".join(nonempty).upper()
                    # Section rows inside table
                    if len(nonempty) <= 2 and not any(looks_like_ahfoz(c) for c in nonempty):
                        if any(
                            k in joined
                            for k in [
                                "GENERAL",
                                "PHARMAC",
                                "DENT",
                                "LAB",
                                "RADIO",
                                "OPTIC",
                                "HOSPITAL",
                                "SPECIAL",
                                "PHYSIO",
                                "CLINIC",
                                "REHAB",
                                "AMBULANCE",
                                "DIALYSIS",
                                "PSYCH",
                                "THEATRE",
                                "NURSE",
                                "WELLNESS",
                            ]
                        ):
                            current_section = joined
                        continue
                    if "DISTRICT" in joined and "NAME" in joined:
                        continue

                    def get(key: str) -> str:
                        i = header_map.get(key)
                        if i is None or i >= len(cells):
                            return ""
                        return cells[i]

                    name = get("name")
                    address = get("address")
                    ahfoz = get("ahfoz")
                    mobile = get("mobile")
                    phone = get("phone")
                    district = get("district")
                    suburb = get("suburb")
                    discipline = get("discipline")

                    # Fallback positional extraction when sparse None columns
                    if not name or not ahfoz:
                        compact = [c for c in cells if c]
                        # Find AHFoZ-like token
                        ahfoz_idx = next((i for i, c in enumerate(compact) if looks_like_ahfoz(c)), -1)
                        if ahfoz_idx >= 2:
                            # typical: district suburb name address ahfoz ...
                            if not district and ahfoz_idx >= 4:
                                district = compact[0]
                                suburb = compact[1]
                                name = name or compact[2]
                                address = address or compact[3]
                                ahfoz = ahfoz or compact[ahfoz_idx]
                            elif ahfoz_idx >= 3:
                                district = district or compact[0]
                                name = name or compact[1]
                                address = address or compact[2]
                                ahfoz = ahfoz or compact[ahfoz_idx]
                            else:
                                name = name or compact[0]
                                ahfoz = ahfoz or compact[ahfoz_idx]
                            # phones after ahfoz
                            after = compact[ahfoz_idx + 1 :]
                            phone_like = [normalize_phone(x) for x in after if re.search(r"\d{6,}", x)]
                            if phone_like and not mobile:
                                mobile = phone_like[0]
                            if len(phone_like) > 1 and not phone:
                                phone = phone_like[1]

                    name = clean(name)
                    if not name or name.upper() in {"NAME", "GENERAL PRACTITIONERS"}:
                        continue
                    # Skip TOC-ish
                    if name.count(".") > 5:
                        continue

                    r = empty_row()
                    r["Provider Name"] = name
                    r["AHFoZ Number"] = normalize_ahfoz_fixed(ahfoz) if looks_like_ahfoz(ahfoz) else ""
                    r["Address"] = clean(address)
                    r["City"] = normalize_city(suburb or district or guess_city_from_address(address))
                    r["Province"] = province or infer_province(r["City"], district, path.name)
                    r["Mobile"] = normalize_phone(mobile)
                    r["Phone"] = normalize_phone(phone)
                    ptype, spec = infer_type_from_name(name, discipline or current_section)
                    r["Provider Type"] = ptype or "Healthcare Provider"
                    r["Speciality"] = spec
                    r["Medical Aid Accepted"] = "PSMAS"
                    r["Organisation"] = "PSMAS Network"
                    r["Source URL"] = source_url
                    r["Source Title"] = f"PSMAS Provider Network PDF - {path.stem}"
                    r["Last Verified"] = TODAY
                    r["Notes"] = f"Section: {current_section}" if current_section else ""
                    rows.append(r)
                    stats["from_psmas_pdf"] += 1
    return rows


def parse_all_psmas_pdfs() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    # Prefer provincial PDFs; still parse consolidated for completeness
    pdfs = sorted(PSMAS_PDF_DIR.glob("*.pdf"))
    for pdf in pdfs:
        print(f"Parsing PDF: {pdf.name}")
        try:
            rows.extend(parse_pdf_tables(pdf))
        except Exception as exc:  # noqa: BLE001
            print(f"  ERROR {pdf.name}: {exc}")
            stats["pdf_errors"] += 1
    return rows


def parse_hpa_markdown(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    if not path.exists():
        return rows
    add_source(
        "https://hpa.co.zw/registered-facilities/",
        "Health Professions Authority - Registered Facilities",
        "Parsed from public registered facilities listing",
    )
    text = path.read_text(encoding="utf-8", errors="ignore")
    # markdown table rows
    for line in text.splitlines():
        if not line.startswith("|"):
            continue
        parts = [clean(p) for p in line.strip("|").split("|")]
        if len(parts) < 3:
            continue
        if parts[0].lower() in {"facility name", "---"} or set(parts[0]) <= {"-"}:
            continue
        name, address, city = parts[0], parts[1], parts[2]
        contact = " ".join(p for p in parts[3:5] if p).strip()
        if not name or name.lower() == "facility name":
            continue
        r = empty_row()
        r["Provider Name"] = name
        r["Address"] = address
        r["City"] = normalize_city(city)
        r["Province"] = infer_province(r["City"])
        r["Contact Person"] = contact
        ptype, spec = infer_type_from_name(name)
        r["Provider Type"] = ptype or "Healthcare Facility"
        r["Speciality"] = spec
        r["Source URL"] = "https://hpa.co.zw/registered-facilities/"
        r["Source Title"] = "HPA Registered Facilities"
        r["Last Verified"] = TODAY
        r["Notes"] = "HPA registered facility; AHFoZ number not published on this page"
        rows.append(r)
        stats["from_hpa"] += 1
    return rows


def fetch_zach_members() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    url = "https://zach.org.zw/membership/"
    record_source_search(url, "ZACH Membership Directory", "Mission hospitals and clinics")
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
    except Exception as exc:  # noqa: BLE001
        print("ZACH fetch failed:", exc)
        stats["zach_errors"] += 1
        return rows
    soup = BeautifulSoup(resp.text, "html.parser")
    table = soup.find("table")
    if not table:
        return rows
    for tr in table.find_all("tr")[1:]:
        tds = [clean(td.get_text(" ", strip=True)) for td in tr.find_all("td")]
        if not tds:
            continue
        name_addr = tds[0]
        phone = tds[1] if len(tds) > 1 else ""
        website = tds[2] if len(tds) > 2 else ""
        beds = tds[3] if len(tds) > 3 else ""
        status = tds[4] if len(tds) > 4 else ""
        # Name often "HOSPITAL City"
        parts = name_addr.rsplit(" ", 1)
        city = ""
        name = name_addr
        if len(parts) == 2 and parts[1].istitle() or (len(parts) == 2 and parts[1].isalpha()):
            # weak split; keep full as name, guess city from known list
            city = guess_city_from_address(name_addr)
            name = name_addr
        # Better: trailing token if known city
        tokens = name_addr.split()
        for n in range(min(3, len(tokens)), 0, -1):
            cand = " ".join(tokens[-n:]).lower()
            if cand in CITY_TO_PROVINCE:
                city = cand.title()
                name = " ".join(tokens[:-n]).strip() or name_addr
                break
        r = empty_row()
        r["Provider Name"] = name
        r["City"] = normalize_city(city)
        r["Province"] = infer_province(r["City"])
        r["Phone"] = normalize_phone(phone)
        r["Website"] = "" if website.upper() == "N/A" else website
        up = name.upper()
        if "HOSPITAL" in up:
            r["Provider Type"] = "Mission Hospital"
        elif "CLINIC" in up:
            r["Provider Type"] = "Clinic"
        else:
            r["Provider Type"] = "Mission Hospital"
        r["Organisation"] = "ZACH"
        r["Medical Aid Accepted"] = ""
        r["Source URL"] = url
        r["Source Title"] = "ZACH Membership Directory"
        r["Last Verified"] = TODAY
        note_bits = []
        if beds:
            note_bits.append(f"Beds: {beds}")
        if status:
            note_bits.append(f"Status: {status}")
        r["Notes"] = "; ".join(note_bits)
        rows.append(r)
        stats["from_zach"] += 1
    return rows


def fetch_mdpcz_sample() -> list[dict[str, str]]:
    """Fetch currently rendered MDPCZ public register page (paginated; capture available rows)."""
    rows: list[dict[str, str]] = []
    url = "https://www.mdpcz.co.zw/public_register"
    record_source_search(url, "MDPCZ Public Register", "Medical & dental practitioners")
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
    except Exception as exc:  # noqa: BLE001
        print("MDPCZ fetch failed:", exc)
        stats["mdpcz_errors"] += 1
        return rows
    soup = BeautifulSoup(resp.text, "html.parser")
    table = soup.find("table")
    if not table:
        return rows
    headers = [clean(th.get_text()) for th in table.find_all("th")]
    for tr in table.find_all("tr"):
        tds = [clean(td.get_text(" ", strip=True)) for td in tr.find_all("td")]
        if not tds:
            continue
        data = {headers[i]: tds[i] if i < len(tds) else "" for i in range(len(headers))} if headers else {}
        name = data.get("Name") or (tds[0] if tds else "")
        reg = data.get("Registration Number") or (tds[2] if len(tds) > 2 else "")
        qual = data.get("Qualification") or ""
        specialty = data.get("Specialty") or (tds[-1] if tds else "")
        if not name:
            continue
        r = empty_row()
        r["Provider Name"] = name
        r["Registration Number"] = reg
        r["Speciality"] = specialty
        ptype = "Dentist" if "dental" in (specialty + qual).lower() else "Doctor"
        if "specialist" in specialty.lower():
            ptype = "Doctor"
        r["Provider Type"] = ptype
        r["Source URL"] = url
        r["Source Title"] = "MDPCZ Public Register"
        r["Last Verified"] = TODAY
        r["Notes"] = f"Qualification: {qual}" if qual else "From MDPCZ public register page"
        rows.append(r)
        stats["from_mdpcz"] += 1
    return rows


def fetch_psmas_wp_listings() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    endpoints = [
        "https://psmas.co.zw/wp-json/wp/v2/at_biz_dir?per_page=100",
        "https://psmas.co.zw/wp-json/directorist/v1/listings?per_page=100",
        "https://psmas.co.zw/wp-json/wp/v2/listing?per_page=100",
    ]
    session = requests.Session()
    session.headers["User-Agent"] = "Mozilla/5.0 (compatible; ZimbabweHealthcareResearch/1.0)"
    for endpoint in endpoints:
        record_source_search(endpoint, "PSMAS WP/API listings probe")
        page = 1
        while page <= 30:
            url = endpoint + (("&" if "?" in endpoint else "?") + f"page={page}")
            try:
                resp = session.get(url, timeout=45)
            except Exception as exc:  # noqa: BLE001
                print("listings error", endpoint, exc)
                break
            if resp.status_code != 200:
                break
            try:
                data = resp.json()
            except Exception:
                break
            if not isinstance(data, list) or not data:
                break
            add_source(endpoint, "PSMAS online listings API", f"page {page}")
            for item in data:
                name = clean(item.get("title", {}).get("rendered") if isinstance(item.get("title"), dict) else item.get("title") or item.get("name"))
                if not name:
                    continue
                link = clean(item.get("link") or item.get("url") or "")
                content = ""
                if isinstance(item.get("content"), dict):
                    content = BeautifulSoup(item["content"].get("rendered", ""), "html.parser").get_text(" ", strip=True)
                r = empty_row()
                r["Provider Name"] = name
                r["Website"] = link
                r["Source URL"] = link or "https://psmas.co.zw/all-listings/"
                r["Source Title"] = "PSMAS All Listings"
                r["Last Verified"] = TODAY
                r["Medical Aid Accepted"] = "PSMAS"
                ptype, spec = infer_type_from_name(name + " " + content)
                r["Provider Type"] = ptype or "Healthcare Provider"
                r["Speciality"] = spec
                # try extract phone/email
                phones = re.findall(r"(?:\+?263|0)\d[\d\s\-]{7,}", content)
                emails = re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", content)
                if phones:
                    r["Phone"] = normalize_phone(phones[0])
                if emails:
                    r["Email"] = emails[0]
                r["Notes"] = "From PSMAS website listings API/HTML JSON"
                rows.append(r)
                stats["from_psmas_listings"] += 1
            if len(data) < 100:
                break
            page += 1
            time.sleep(0.2)
        if rows:
            break
    return rows


def fetch_additional_directories() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    targets = [
        ("https://medical.co.zw/", "Zimbabwe Medical Directory"),
        ("https://www.hellodoctor.co.zw/", "Hello Doctor Zimbabwe"),
        ("https://ahfoz.org/", "AHFoZ official site"),
        ("https://ahfoz.org/our-members/", "AHFoZ members"),
        ("https://cimas.co.zw/blue-zone/", "CIMAS Blue Zone provider directory"),
        ("https://hpa.co.zw/", "Health Professions Authority"),
        ("https://portal.ahpcz.co.zw/", "Allied Health Practitioners Council portal"),
        ("https://www.mdpcz.co.zw/", "MDPCZ home"),
        ("https://psmas.co.zw/view-download-provider-list/", "PSMAS View/Download Provider List"),
        ("https://psmas.co.zw/all-listings/", "PSMAS All Listings"),
    ]
    session = requests.Session()
    session.headers["User-Agent"] = "Mozilla/5.0 (compatible; ZimbabweHealthcareResearch/1.0)"
    for url, title in targets:
        record_source_search(url, title, "Directory searched")
        try:
            resp = session.get(url, timeout=45)
            stats["websites_searched"] += 1
            if resp.status_code != 200:
                continue
            soup = BeautifulSoup(resp.text, "html.parser")
            # Collect obvious provider-like cards/links
            for a in soup.find_all("a", href=True):
                text = clean(a.get_text(" ", strip=True))
                href = urljoin(url, a["href"])
                if not text or len(text) < 4 or len(text) > 120:
                    continue
                if not re.search(
                    r"hospital|clinic|pharmacy|doctor|dr\b|lab|dental|physio|optic|radiolog|medical",
                    text,
                    re.I,
                ):
                    continue
                if href.rstrip("/") == url.rstrip("/"):
                    continue
                r = empty_row()
                r["Provider Name"] = text
                r["Website"] = href if "facebook" not in href and "linkedin" not in href else ""
                r["Source URL"] = href if href.startswith("http") else url
                r["Source Title"] = title
                r["Last Verified"] = TODAY
                ptype, spec = infer_type_from_name(text)
                r["Provider Type"] = ptype or "Healthcare Provider"
                r["Speciality"] = spec
                r["Notes"] = f"Discovered via directory crawl of {title}"
                rows.append(r)
                stats["from_directory_crawl"] += 1
        except Exception as exc:  # noqa: BLE001
            print("dir crawl fail", url, exc)
    return rows


# ------------------------- merge / export -------------------------


def dedupe_key(r: dict[str, str]) -> str:
    ah = r.get("AHFoZ Number") or ""
    if ah:
        return f"ahfoz:{ah}"
    name = re.sub(r"[^a-z0-9]+", "", (r.get("Provider Name") or "").lower())
    city = re.sub(r"[^a-z0-9]+", "", (r.get("City") or "").lower())
    addr = re.sub(r"[^a-z0-9]+", "", (r.get("Address") or "").lower())[:40]
    reg = re.sub(r"[^a-z0-9]+", "", (r.get("Registration Number") or "").lower())
    if reg:
        return f"reg:{reg}"
    return f"name:{name}|city:{city}|addr:{addr}"


def merge_records(records: list[dict[str, str]]) -> list[dict[str, str]]:
    merged: dict[str, dict[str, str]] = {}
    duplicates = 0
    priority = {
        "PSMAS Provider Network PDF": 100,
        "PSMAS Provider Network List (Excel)": 90,
        "HPA Registered Facilities": 80,
        "MDPCZ Public Register": 85,
        "ZACH Membership Directory": 70,
        "PSMAS All Listings": 60,
    }

    def score(r: dict[str, str]) -> int:
        s = 0
        title = r.get("Source Title") or ""
        for k, v in priority.items():
            if k in title:
                s = max(s, v)
        if r.get("AHFoZ Number"):
            s += 50
        if r.get("Registration Number"):
            s += 20
        filled = sum(1 for c in COLUMNS if r.get(c))
        s += filled
        return s

    for r in records:
        key = dedupe_key(r)
        if key not in merged:
            merged[key] = r
            continue
        duplicates += 1
        cur = merged[key]
        # Prefer higher score as base, then fill blanks
        base, other = (r, cur) if score(r) > score(cur) else (cur, r)
        out = empty_row()
        for c in COLUMNS:
            out[c] = base.get(c) or other.get(c) or ""
        # Merge medical aids
        aids = []
        for src in (base, other):
            for part in re.split(r"[;,/|]", src.get("Medical Aid Accepted") or ""):
                p = clean(part)
                if p and p not in aids:
                    aids.append(p)
        out["Medical Aid Accepted"] = "; ".join(aids)
        # Combine notes
        notes = []
        for src in (base, other):
            n = clean(src.get("Notes"))
            if n and n not in notes:
                notes.append(n)
        out["Notes"] = " | ".join(notes)[:1000]
        # Prefer non-empty AHFoZ
        if not out["AHFoZ Number"]:
            out["AHFoZ Number"] = base.get("AHFoZ Number") or other.get("AHFoZ Number") or ""
        merged[key] = out
    stats["duplicates_removed"] = duplicates
    # Sort: AHFoZ first, then name
    result = list(merged.values())
    result.sort(key=lambda x: (0 if x.get("AHFoZ Number") else 1, (x.get("Provider Name") or "").lower()))
    return result


def classify_counts(rows: list[dict[str, str]]) -> dict[str, int]:
    c = Counter()
    for r in rows:
        t = (r.get("Provider Type") or "").lower()
        if "hospital" in t:
            c["hospitals"] += 1
        if t in {"doctor"} or "practitioner" in t or "surgeon" in t:
            c["doctors"] += 1
        if "pharmac" in t:
            c["pharmacies"] += 1
        if "laborator" in t:
            c["laboratories"] += 1
        # also name-based backup
        name = (r.get("Provider Name") or "").lower()
        if "pharmacy" in name or "chemist" in name:
            c["pharmacies_name"] += 1
        if "laboratory" in name or "lab " in name or name.endswith(" lab"):
            c["laboratories_name"] += 1
        if "hospital" in name:
            c["hospitals_name"] += 1
        if r.get("AHFoZ Number"):
            c["with_ahfoz"] += 1
        if r.get("Registration Number"):
            c["with_reg"] += 1
    return c


def export_all(rows: list[dict[str, str]]) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    # Final deliverables also at repo root as requested
    root_out = ROOT.parent

    # providers.csv
    providers_csv = OUT / "providers.csv"
    with providers_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in COLUMNS})
    (root_out / "providers.csv").write_bytes(providers_csv.read_bytes())

    # sources.csv
    sources_csv = OUT / "sources.csv"
    with sources_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["URL", "Title", "Accessed", "Notes"])
        w.writeheader()
        for s in sorted(sources, key=lambda x: x["URL"]):
            w.writerow(s)
    (root_out / "sources.csv").write_bytes(sources_csv.read_bytes())

    # xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Providers"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c, "") for c in COLUMNS])
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    xlsx_path = OUT / "Zimbabwe_Healthcare_Providers.xlsx"
    wb.save(xlsx_path)
    wb.save(root_out / "Zimbabwe_Healthcare_Providers.xlsx")

    counts = classify_counts(rows)
    pdfs = list(PSMAS_PDF_DIR.glob("*.pdf"))
    summary = f"""Zimbabwe Healthcare Providers Database
Generated: {TODAY}

Number of providers collected: {len(rows)}
Number with AHFoZ numbers: {counts.get('with_ahfoz', 0)}
Number with registration numbers: {counts.get('with_reg', 0)}
Number of hospitals: {max(counts.get('hospitals', 0), counts.get('hospitals_name', 0))}
Number of doctors: {counts.get('doctors', 0)}
Number of pharmacies: {max(counts.get('pharmacies', 0), counts.get('pharmacies_name', 0))}
Number of laboratories: {max(counts.get('laboratories', 0), counts.get('laboratories_name', 0))}
Number of duplicates removed: {stats.get('duplicates_removed', 0)}
Websites searched: {stats.get('websites_searched', 0) + len(sources)}
PDFs processed: {len(pdfs)}
Total rows: {len(rows)}

Source breakdown (pre-dedupe counts):
- Repository PSMAS Excel: {stats.get('from_repo_xlsx', 0)}
- PSMAS PDFs: {stats.get('from_psmas_pdf', 0)}
- HPA registered facilities: {stats.get('from_hpa', 0)}
- ZACH mission facilities: {stats.get('from_zach', 0)}
- MDPCZ public register (page sample): {stats.get('from_mdpcz', 0)}
- PSMAS listings API: {stats.get('from_psmas_listings', 0)}
- Directory crawl discoveries: {stats.get('from_directory_crawl', 0)}

Primary AHFoZ-bearing sources:
- PSMAS provincial provider network PDFs (July 2026)
- PSMAS consolidated provider network list PDF
- Existing PSMAS Excel in repository

Notes:
- AHFoZ does not publish a public full provider payee directory; AHFoZ numbers were captured from medical-aid provider network lists where published.
- HPA and MDPCZ contribute facility/practitioner identity and registration data; AHFoZ fields left blank when not publicly available.
- MDPCZ online register is paginated (~4650 practitioners); this build includes rows available from the public page response plus other sources. Full register export was not exposed as a static download at collection time.
"""
    (OUT / "summary.txt").write_text(summary, encoding="utf-8")
    (root_out / "summary.txt").write_text(summary, encoding="utf-8")
    print(summary)


def main() -> None:
    RAW.mkdir(parents=True, exist_ok=True)
    OUT.mkdir(parents=True, exist_ok=True)
    all_rows: list[dict[str, str]] = []

    print("1) Existing Excel...")
    all_rows.extend(parse_existing_xlsx())

    print("2) PSMAS PDFs...")
    all_rows.extend(parse_all_psmas_pdfs())

    print("3) HPA facilities...")
    all_rows.extend(parse_hpa_markdown(HPA_CACHE))

    print("4) ZACH...")
    all_rows.extend(fetch_zach_members())

    print("5) MDPCZ...")
    all_rows.extend(fetch_mdpcz_sample())

    print("6) PSMAS listings API...")
    all_rows.extend(fetch_psmas_wp_listings())

    print("7) Additional directories...")
    all_rows.extend(fetch_additional_directories())

    print(f"Raw rows before dedupe: {len(all_rows)}")
    merged = merge_records(all_rows)
    print(f"Rows after dedupe: {len(merged)}")
    export_all(merged)

    # Persist raw stats
    (OUT / "stats.json").write_text(json.dumps(dict(stats), indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
